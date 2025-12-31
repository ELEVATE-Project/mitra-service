from pdf2image import convert_from_path
from docx import Document
from openpyxl import load_workbook
import markdown
import csv
from .base import BasePreviewGenerator


class PDFPreviewGenerator(BasePreviewGenerator):
    """Generate preview from PDF first page"""

    def extract_content(self):
        return convert_from_path(self.file_path, first_page=1, last_page=1)[0]

    def create_image(self, content):
        return content  # PDF already returns an image


class DocxPreviewGenerator(BasePreviewGenerator):
    """Generate preview from DOCX document"""

    def extract_content(self):
        doc = Document(self.file_path)
        return "\n".join(p.text for p in doc.paragraphs[:8])


class XlsxPreviewGenerator(BasePreviewGenerator):
    """Generate preview from Excel spreadsheet"""

    def extract_content(self):
        wb = load_workbook(self.file_path, data_only=True)
        sheet = wb.active

        lines = []
        for row in sheet.iter_rows(max_row=10, max_col=5, values_only=True):
            line = " | ".join(str(cell or "") for cell in row)
            lines.append(line)

        return "\n".join(lines)


class MarkdownPreviewGenerator(BasePreviewGenerator):
    """Generate preview from Markdown file"""

    def extract_content(self):
        with open(self.file_path, 'r', encoding='utf-8') as f:
            content = f.read()

        # Convert markdown to plain text (strip HTML tags)
        html = markdown.markdown(content)
        text = html.replace("<", "").replace(">", "")
        return text


class CSVPreviewGenerator(BasePreviewGenerator):
    """Generate preview from CSV file"""

    def extract_content(self):
        lines = []
        try:
            with open(self.file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for i, row in enumerate(reader):
                    if i >= 10:  # Limit to first 10 rows
                        break
                    line = " | ".join(str(cell) for cell in row[:5])  # First 5 columns
                    lines.append(line)
        except Exception as e:
            lines.append(f"Error reading CSV: {str(e)}")

        return "\n".join(lines)


class TxtPreviewGenerator(BasePreviewGenerator):
    """Generate preview from plain text file"""

    def extract_content(self):
        try:
            with open(self.file_path, 'r', encoding='utf-8') as f:
                return f.read()
        except UnicodeDecodeError:
            # Try with a different encoding if UTF-8 fails
            with open(self.file_path, 'r', encoding='latin-1') as f:
                return f.read()
