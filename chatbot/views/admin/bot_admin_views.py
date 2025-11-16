"""
Custom Import/Export Views for CompanyBot with inline models
Add this to your views.py
"""
import json
import csv
import io
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.db import transaction
from openpyxl import Workbook, load_workbook
from chatbot.models import CompanyBot, Voice, CompanyStateMachine, Company, Profile, ProfileType


def generate_template(format_type):
    """Generate empty template files for import"""
    if format_type == 'json':
        template_data = [{
            "name": "Example Bot",
            "company_slug": "company-slug",
            "context": "Bot context here",
            "max_token": 1000,
            "provider": "openai",
            "bot_type": "qa",
            "voices": [
                {
                    "type": "greeting",
                    "provider": "elevenlabs",
                    "name": "Voice Name",
                    "language": "en"
                }
            ],
            "state_machines": [
                {
                    "name": "Step 1",
                    "step": 1,
                    "bot_question": "What is your name?",
                    "completion_criteria": "User provides name"
                }
            ]
        }]

        response = HttpResponse(
            json.dumps(template_data, indent=2),
            content_type='application/json'
        )
        response['Content-Disposition'] = 'attachment; filename="companybot_template.json"'
        return response

    elif format_type == 'xlsx':
        wb = Workbook()

        # Bots sheet
        ws_bots = wb.active
        ws_bots.title = "Bots"
        ws_bots.append(['bot_id', 'name', 'company_slug', 'context', 'max_token',
                        'provider', 'bot_type'])
        ws_bots.append([1, 'Example Bot', 'company-slug', 'Bot context', 1000,
                        'openai', 'qa'])

        # Voices sheet
        ws_voices = wb.create_sheet("Voices")
        ws_voices.append(['bot_id', 'type', 'provider', 'name', 'language'])
        ws_voices.append([1, 'greeting', 'elevenlabs', 'Voice Name', 'en'])

        # State Machines sheet
        ws_sm = wb.create_sheet("StateMachines")
        ws_sm.append(['bot_id', 'name', 'step', 'bot_question', 'completion_criteria'])
        ws_sm.append([1, 'Step 1', 1, 'What is your name?', 'User provides name'])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="companybot_template.xlsx"'
        return response

    elif format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="companybot_template.csv"'

        writer = csv.writer(response)
        writer.writerow(['name', 'company_slug', 'context', 'bot_type', 'voices_json',
                         'state_machines_json'])
        writer.writerow([
            'Example Bot',
            'company-slug',
            'Bot context',
            'qa',
            '[{"type": "greeting", "provider": "elevenlabs", "name": "Voice Name"}]',
            '[{"name": "Step 1", "step": 1, "bot_question": "What is your name?"}]'
        ])

        return response

@staff_member_required
def export_bots(request):
    """Export CompanyBots with their inline models"""
    user_email = request.user.email
    profile = Profile.objects.filter(email=user_email).first()

    # Check if this is a template request
    is_template = request.GET.get('template', 'false').lower() == 'true'

    # Filter bots based on user permissions
    if request.user.is_superuser:
        bots = CompanyBot.objects.all()
    elif profile and profile.profile_type == ProfileType.MODERATOR:
        bots = CompanyBot.objects.filter(company=profile.company)
    else:
        messages.error(request, "You don't have permission to export bots.")
        return redirect('admin:chatbot_companybot_changelist')

    # Get selected bot IDs if any
    bot_ids = request.GET.get('ids', '').split(',')
    bot_ids = [id for id in bot_ids if id]  # Filter empty strings
    if bot_ids:
        bots = bots.filter(id__in=bot_ids)

    # If no format specified, show format selection page
    export_format = request.GET.get('format')
    if not export_format:
        return render(request, 'admin/export_format.html', {
            'bot_count': bots.count(),
            'selected_ids': ','.join(bot_ids),
        })

    # Generate templates if requested
    if is_template:
        return generate_template(export_format)

    # Export based on format
    if export_format == 'json':
        return export_bots_json(bots)
    elif export_format == 'xlsx':
        return export_bots_xlsx(bots)
    elif export_format == 'csv':
        return export_bots_csv(bots)
    else:
        messages.error(request, "Invalid export format.")
        return redirect('admin:chatbot_companybot_changelist')


def export_bots_json(bots):
    """Export bots as JSON"""
    data = []
    for bot in bots.select_related('company'):
        bot_data = {
            'name': bot.name,
            'company_slug': bot.company.slug,
            'context': bot.context,
            'max_token': bot.max_token,
            'provider': bot.provider,
            'provider_keys': bot.provider_keys,
            'bot_temperature': bot.bot_temperature,
            'top_k': bot.top_k,
            'llm_model': bot.llm_model,
            'filter_score': bot.filter_score,
            'end_context': bot.end_context,
            'introductory_message': bot.introductory_message,
            'tag_context': bot.tag_context,
            'route': bot.route,
            'bot_type': bot.bot_type,
            'llm_key': bot.llm_key,
            'dynamic_context': bot.dynamic_context,
            'dynamic_context_type': bot.dynamic_context_type,
            'pre_context': bot.pre_context,
            'tool_context': bot.tool_context,
            'other_params': bot.other_params,
            'connect_timeout': bot.connect_timeout,
            'read_timeout': bot.read_timeout,
        }

        # Add voices
        voices = Voice.objects.filter(company_bot=bot)
        bot_data['voices'] = [
            {
                'type': v.type,
                'provider': v.provider,
                'name': v.name,
                'sample_link': v.sample_link,
                'language': v.language,
                'provider_code': v.provider_code,
                'gender': v.gender,
                'voice_speed': v.voice_speed,
                'other_params': v.other_params,
            }
            for v in voices
        ]

        # Add state machines
        state_machines = CompanyStateMachine.objects.filter(company_bot=bot).order_by('step')
        bot_data['state_machines'] = [
            {
                'name': sm.name,
                'step': sm.step,
                'use_stage_chats': sm.use_stage_chats,
                'type': sm.type,
                'text_conversion_type': sm.text_conversion_type,
                'bot_question': sm.bot_question,
                'completion_criteria': sm.completion_criteria,
                'context': sm.context,
                'tool_context': sm.tool_context,
                'preprocess_type': sm.preprocess_type,
                'preprocess_prompt': sm.preprocess_prompt,
                'preprocess_bot_name': sm.preprocess_bot.name if sm.preprocess_bot else None,
                'preprocess_output_mode': sm.preprocess_output_mode,
                'postprocess_type': sm.postprocess_type,
                'postprocess_prompt': sm.postprocess_prompt,
                'postprocess_bot_name': sm.postprocess_bot.name if sm.postprocess_bot else None,
                'postprocess_output_mode': sm.postprocess_output_mode,
                'skip_to_step': sm.skip_to_step,
            }
            for sm in state_machines
        ]

        data.append(bot_data)

    response = HttpResponse(
        json.dumps(data, indent=2, default=str),
        content_type='application/json'
    )
    response['Content-Disposition'] = 'attachment; filename="company_bots.json"'
    return response


def export_bots_xlsx(bots):
    """Export bots as Excel file with multiple sheets"""
    wb = Workbook()

    # Main bot sheet
    ws_bots = wb.active
    ws_bots.title = "Bots"
    bot_headers = [
        'bot_id', 'name', 'company_slug', 'context', 'max_token', 'provider',
        'provider_keys', 'bot_temperature', 'top_k', 'llm_model', 'filter_score',
        'end_context', 'introductory_message', 'tag_context', 'route', 'bot_type',
        'llm_key', 'dynamic_context', 'dynamic_context_type', 'pre_context',
        'tool_context', 'other_params', 'connect_timeout', 'read_timeout'
    ]
    ws_bots.append(bot_headers)

    # Voice sheet
    ws_voices = wb.create_sheet("Voices")
    voice_headers = [
        'bot_id', 'type', 'provider', 'name', 'sample_link', 'language',
        'provider_code', 'gender', 'voice_speed', 'other_params'
    ]
    ws_voices.append(voice_headers)

    # State machine sheet
    ws_sm = wb.create_sheet("StateMachines")
    sm_headers = [
        'bot_id', 'name', 'step', 'use_stage_chats', 'type', 'text_conversion_type',
        'bot_question', 'completion_criteria', 'context', 'tool_context',
        'preprocess_type', 'preprocess_prompt', 'preprocess_bot_name',
        'preprocess_output_mode', 'postprocess_type', 'postprocess_prompt',
        'postprocess_bot_name', 'postprocess_output_mode', 'skip_to_step'
    ]
    ws_sm.append(sm_headers)

    # Populate data
    for bot in bots.select_related('company'):
        # Bot row
        bot_row = [
            bot.id, bot.name, bot.company.slug, bot.context, bot.max_token,
            bot.provider, bot.provider_keys, bot.bot_temperature, bot.top_k,
            bot.llm_model, bot.filter_score, bot.end_context,
            bot.introductory_message, bot.tag_context, bot.route, bot.bot_type,
            bot.llm_key, bot.dynamic_context, bot.dynamic_context_type,
            bot.pre_context, bot.tool_context,
            json.dumps(bot.other_params) if bot.other_params else None,
            bot.connect_timeout, bot.read_timeout
        ]
        ws_bots.append(bot_row)

        # Voice rows
        for voice in Voice.objects.filter(company_bot=bot):
            voice_row = [
                bot.id, voice.type, voice.provider, voice.name, voice.sample_link,
                voice.language, voice.provider_code, voice.gender, voice.voice_speed,
                json.dumps(voice.other_params) if voice.other_params else None
            ]
            ws_voices.append(voice_row)

        # State machine rows
        for sm in CompanyStateMachine.objects.filter(company_bot=bot).order_by('step'):
            sm_row = [
                bot.id, sm.name, sm.step, sm.use_stage_chats, sm.type,
                sm.text_conversion_type, sm.bot_question, sm.completion_criteria,
                sm.context, sm.tool_context, sm.preprocess_type, sm.preprocess_prompt,
                sm.preprocess_bot.name if sm.preprocess_bot else None,
                sm.preprocess_output_mode, sm.postprocess_type, sm.postprocess_prompt,
                sm.postprocess_bot.name if sm.postprocess_bot else None,
                sm.postprocess_output_mode, sm.skip_to_step
            ]
            ws_sm.append(sm_row)

    # Save to response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="company_bots.xlsx"'
    return response


def export_bots_csv(bots):
    """Export bots as CSV (flattened structure)"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="company_bots.csv"'

    writer = csv.writer(response)
    headers = [
        'name', 'company_slug', 'context', 'max_token', 'provider',
        'bot_temperature', 'top_k', 'llm_model', 'filter_score',
        'end_context', 'introductory_message', 'route', 'bot_type',
        'voices_json', 'state_machines_json'
    ]
    writer.writerow(headers)

    for bot in bots.select_related('company'):
        # Serialize voices
        voices = Voice.objects.filter(company_bot=bot)
        voices_data = [
            {
                'type': v.type,
                'provider': v.provider,
                'name': v.name,
                'language': v.language,
                'provider_code': v.provider_code,
                'gender': v.gender,
            }
            for v in voices
        ]

        # Serialize state machines
        state_machines = CompanyStateMachine.objects.filter(company_bot=bot).order_by('step')
        sm_data = [
            {
                'name': sm.name,
                'step': sm.step,
                'bot_question': sm.bot_question,
                'completion_criteria': sm.completion_criteria,
            }
            for sm in state_machines
        ]

        row = [
            bot.name, bot.company.slug, bot.context, bot.max_token, bot.provider,
            bot.bot_temperature, bot.top_k, bot.llm_model, bot.filter_score,
            bot.end_context, bot.introductory_message, bot.route, bot.bot_type,
            json.dumps(voices_data), json.dumps(sm_data)
        ]
        writer.writerow(row)

    return response


@staff_member_required
def import_bots(request):
    """Import CompanyBots with their inline models"""
    if request.method == 'POST':
        uploaded_file = request.FILES.get('import_file')
        if not uploaded_file:
            messages.error(request, "Please upload a file.")
            return redirect('admin:chatbot_companybot_changelist')

        file_extension = uploaded_file.name.split('.')[-1].lower()

        try:
            if file_extension == 'json':
                result = import_bots_json(request, uploaded_file)
            elif file_extension == 'xlsx':
                result = import_bots_xlsx(request, uploaded_file)
            elif file_extension == 'csv':
                result = import_bots_csv(request, uploaded_file)
            else:
                messages.error(request, "Unsupported file format. Use JSON, XLSX, or CSV.")
                return redirect('admin:chatbot_companybot_changelist')

            messages.success(
                request,
                f"Successfully imported {result['created']} bots and updated {result['updated']} bots."
            )
        except Exception as e:
            messages.error(request, f"Import failed: {str(e)}")

        return redirect('admin:chatbot_companybot_changelist')

    # GET request - show import form
    return render(request, "admin/import_form.html")


def import_bots_json(request, uploaded_file):
    """Import from JSON file"""
    data = json.load(uploaded_file)
    user_email = request.user.email
    profile = Profile.objects.filter(email=user_email).first()

    created_count = 0
    updated_count = 0

    with transaction.atomic():
        for bot_data in data:
            company_slug = bot_data.pop('company_slug')
            try:
                company = Company.objects.get(slug=company_slug)
            except Company.DoesNotExist:
                raise ValueError(f"Company with slug '{company_slug}' not found")

            # Check permissions
            if not request.user.is_superuser:
                if not profile or profile.profile_type != ProfileType.MODERATOR or profile.company != company:
                    raise PermissionError(f"You don't have permission to import bots for company {company_slug}")

            # Extract inline data
            voices_data = bot_data.pop('voices', [])
            state_machines_data = bot_data.pop('state_machines', [])

            # Create or update bot
            bot, created = CompanyBot.objects.update_or_create(
                name=bot_data['name'],
                company=company,
                defaults=bot_data
            )

            if created:
                created_count += 1
            else:
                updated_count += 1

            # Delete existing inline records
            Voice.objects.filter(company_bot=bot).delete()
            CompanyStateMachine.objects.filter(company_bot=bot).delete()

            # Create voices
            for voice_data in voices_data:
                Voice.objects.create(company_bot=bot, **voice_data)

            # Create state machines
            for sm_data in state_machines_data:
                # Handle bot references
                preprocess_bot_name = sm_data.pop('preprocess_bot_name', None)
                postprocess_bot_name = sm_data.pop('postprocess_bot_name', None)

                preprocess_bot = None
                if preprocess_bot_name:
                    try:
                        preprocess_bot = CompanyBot.objects.get(name=preprocess_bot_name, company=company)
                    except CompanyBot.DoesNotExist:
                        pass

                postprocess_bot = None
                if postprocess_bot_name:
                    try:
                        postprocess_bot = CompanyBot.objects.get(name=postprocess_bot_name, company=company)
                    except CompanyBot.DoesNotExist:
                        pass

                CompanyStateMachine.objects.create(
                    company_bot=bot,
                    preprocess_bot=preprocess_bot,
                    postprocess_bot=postprocess_bot,
                    **sm_data
                )

    return {'created': created_count, 'updated': updated_count}


def import_bots_xlsx(request, uploaded_file):
    """Import from Excel file"""
    wb = load_workbook(uploaded_file)
    user_email = request.user.email
    profile = Profile.objects.filter(email=user_email).first()

    created_count = 0
    updated_count = 0

    # Read sheets
    ws_bots = wb['Bots']
    ws_voices = wb.get('Voices')
    ws_sm = wb.get('StateMachines')

    # Parse headers
    bot_headers = [cell.value for cell in ws_bots[1]]

    with transaction.atomic():
        # Process bots
        for row in ws_bots.iter_rows(min_row=2, values_only=True):
            bot_data = dict(zip(bot_headers, row))
            bot_id = bot_data.pop('bot_id', None)
            company_slug = bot_data.pop('company_slug')

            try:
                company = Company.objects.get(slug=company_slug)
            except Company.DoesNotExist:
                continue

            # Check permissions
            if not request.user.is_superuser:
                if not profile or profile.profile_type != ProfileType.MODERATOR or profile.company != company:
                    continue

            # Parse JSON fields
            if bot_data.get('other_params'):
                bot_data['other_params'] = json.loads(bot_data['other_params'])

            # Create or update bot
            bot, created = CompanyBot.objects.update_or_create(
                name=bot_data['name'],
                company=company,
                defaults={k: v for k, v in bot_data.items() if v is not None and k != 'name'}
            )

            if created:
                created_count += 1
            else:
                updated_count += 1
                # Delete existing inline records for updates
                Voice.objects.filter(company_bot=bot).delete()
                CompanyStateMachine.objects.filter(company_bot=bot).delete()

            # Process voices
            if ws_voices:
                voice_headers = [cell.value for cell in ws_voices[1]]
                for row in ws_voices.iter_rows(min_row=2, values_only=True):
                    voice_data = dict(zip(voice_headers, row))
                    if voice_data.get('bot_id') == bot_id or (not bot_id and voice_data.get('bot_id') == bot.id):
                        voice_data.pop('bot_id')
                        if voice_data.get('other_params'):
                            voice_data['other_params'] = json.loads(voice_data['other_params'])
                        Voice.objects.create(company_bot=bot, **{k: v for k, v in voice_data.items() if v is not None})

            # Process state machines
            if ws_sm:
                sm_headers = [cell.value for cell in ws_sm[1]]
                for row in ws_sm.iter_rows(min_row=2, values_only=True):
                    sm_data = dict(zip(sm_headers, row))
                    if sm_data.get('bot_id') == bot_id or (not bot_id and sm_data.get('bot_id') == bot.id):
                        sm_data.pop('bot_id')

                        # Handle bot references
                        preprocess_bot_name = sm_data.pop('preprocess_bot_name', None)
                        postprocess_bot_name = sm_data.pop('postprocess_bot_name', None)

                        preprocess_bot = None
                        if preprocess_bot_name:
                            try:
                                preprocess_bot = CompanyBot.objects.get(name=preprocess_bot_name, company=company)
                            except CompanyBot.DoesNotExist:
                                pass

                        postprocess_bot = None
                        if postprocess_bot_name:
                            try:
                                postprocess_bot = CompanyBot.objects.get(name=postprocess_bot_name, company=company)
                            except CompanyBot.DoesNotExist:
                                pass

                        CompanyStateMachine.objects.create(
                            company_bot=bot,
                            preprocess_bot=preprocess_bot,
                            postprocess_bot=postprocess_bot,
                            **{k: v for k, v in sm_data.items() if v is not None}
                        )

    return {'created': created_count, 'updated': updated_count}


def import_bots_csv(request, uploaded_file):
    """Import from CSV file"""
    content = uploaded_file.read().decode('utf-8')
    csv_reader = csv.DictReader(io.StringIO(content))

    user_email = request.user.email
    profile = Profile.objects.filter(email=user_email).first()

    created_count = 0
    updated_count = 0

    with transaction.atomic():
        for row in csv_reader:
            company_slug = row.pop('company_slug')
            try:
                company = Company.objects.get(slug=company_slug)
            except Company.DoesNotExist:
                continue

            # Check permissions
            if not request.user.is_superuser:
                if not profile or profile.profile_type != ProfileType.MODERATOR or profile.company != company:
                    continue

            # Extract inline data
            voices_data = json.loads(row.pop('voices_json', '[]'))
            state_machines_data = json.loads(row.pop('state_machines_json', '[]'))

            # Create or update bot
            bot, created = CompanyBot.objects.update_or_create(
                name=row['name'],
                company=company,
                defaults={k: v for k, v in row.items() if v and k != 'name'}
            )

            if created:
                created_count += 1
            else:
                updated_count += 1
                Voice.objects.filter(company_bot=bot).delete()
                CompanyStateMachine.objects.filter(company_bot=bot).delete()

            # Create voices
            for voice_data in voices_data:
                Voice.objects.create(company_bot=bot, **voice_data)

            # Create state machines
            for sm_data in state_machines_data:
                CompanyStateMachine.objects.create(company_bot=bot, **sm_data)

    return {'created': created_count, 'updated': updated_count}